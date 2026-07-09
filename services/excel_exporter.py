import io
import zipfile
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.excel import ClientMergeRange, ClientTableItem, ExportOptions
from utils.excel_utils import sanitize_sheet_name, sanitize_file_name

def build_workbook_for_table_python(table: ClientTableItem, idx: int) -> Workbook:
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
        
    sheet_name = sanitize_sheet_name(table.tableName, f"Sheet {idx + 1}")
    
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="4A6CF7", end_color="4A6CF7", fill_type="solid")
    
    font_body = Font(name="Arial", size=10, color="000000")
    fill_zebra = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D0D0D0")
    medium_bottom_side = Side(border_style="medium", color="4A6CF7")
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=medium_bottom_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    current_row = 1
    
    if table.headerRows is not None:
        for r_idx, r_data in enumerate(table.rows):
            is_header = r_idx in table.headerRows
            ws.row_dimensions[current_row].height = 25 if is_header else 20
            
            for c_idx, val in enumerate(r_data):
                col_letter = get_column_letter(c_idx + 1)
                cell = ws[f"{col_letter}{current_row}"]
                cell.value = val
                cell.alignment = align_center if is_header else align_left
                
                if is_header:
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.border = border_header
                else:
                    cell.font = font_body
                    cell.border = border_cell
                    if r_idx % 2 == 1:
                        cell.fill = fill_zebra
            current_row += 1
    else:
        headers_list = []
        if table.headers:
            if isinstance(table.headers[0], list):
                headers_list = table.headers
            else:
                headers_list = [table.headers]
                
        for h_row in headers_list:
            ws.row_dimensions[current_row].height = 25
            for c_idx, val in enumerate(h_row):
                col_letter = get_column_letter(c_idx + 1)
                cell = ws[f"{col_letter}{current_row}"]
                cell.value = val
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border_header
            current_row += 1
            
        for r_idx, r_data in enumerate(table.rows):
            ws.row_dimensions[current_row].height = 20
            is_zebra = (r_idx % 2 == 1)
            for c_idx, val in enumerate(r_data):
                col_letter = get_column_letter(c_idx + 1)
                cell = ws[f"{col_letter}{current_row}"]
                cell.value = val
                cell.font = font_body
                if is_zebra:
                    cell.fill = fill_zebra
                cell.alignment = align_left
                cell.border = border_cell
            current_row += 1
        
    if table.merges:
        for merge in table.merges:
            try:
                for r in range(merge.startRow + 1, merge.endRow + 2):
                    for c in range(merge.startCol + 1, merge.endCol + 2):
                        ws.cell(row=r, column=c).border = border_cell
                
                ws.merge_cells(
                    start_row=merge.startRow + 1,
                    start_column=merge.startCol + 1,
                    end_row=merge.endRow + 1,
                    end_column=merge.endCol + 1
                )
            except Exception as e:
                pass
                
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ws.merged_cells:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(10, min(50, max_len + 3))
        
    return wb

def build_combined_vertical_workbook(tables: List[ClientTableItem]) -> Workbook:
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
        
    ws = wb.create_sheet(title="Dữ liệu gộp dọc")
    ws.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="Arial", size=14, bold=True, color="1F2937")
    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="4A6CF7", end_color="4A6CF7", fill_type="solid")
    font_body = Font(name="Arial", size=10, color="000000")
    fill_zebra = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="E0E0E0")
    medium_bottom_side = Side(border_style="medium", color="4A6CF7")
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=medium_bottom_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    current_row = 1
    
    for idx, table in enumerate(tables):
        title_text = table.tableName or f"Bảng {idx + 1}"
        ws.cell(row=current_row, column=1, value=title_text).font = font_title
        current_row += 2
        
        row_offset = current_row - 1
        
        if table.headerRows is not None:
            for r_idx, r_data in enumerate(table.rows):
                is_header = r_idx in table.headerRows
                ws.row_dimensions[current_row].height = 25 if is_header else 20
                for c_idx, val in enumerate(r_data):
                    cell = ws.cell(row=current_row, column=c_idx + 1, value=val)
                    cell.alignment = align_center if is_header else align_left
                    if is_header:
                        cell.font = font_header
                        cell.fill = fill_header
                        cell.border = border_header
                    else:
                        cell.font = font_body
                        cell.border = border_cell
                        if r_idx % 2 == 1:
                            cell.fill = fill_zebra
                current_row += 1
        else:
            headers_list = []
            if table.headers:
                if isinstance(table.headers[0], list):
                    headers_list = table.headers
                else:
                    headers_list = [table.headers]
                    
            for h_row in headers_list:
                ws.row_dimensions[current_row].height = 25
                for c_idx, val in enumerate(h_row):
                    cell = ws.cell(row=current_row, column=c_idx + 1, value=val)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = border_header
                current_row += 1
                
            for r_idx, r_data in enumerate(table.rows):
                ws.row_dimensions[current_row].height = 20
                is_zebra = (r_idx % 2 == 1)
                for c_idx, val in enumerate(r_data):
                    cell = ws.cell(row=current_row, column=c_idx + 1, value=val)
                    cell.font = font_body
                    if is_zebra:
                        cell.fill = fill_zebra
                    cell.alignment = align_left
                    cell.border = border_cell
                current_row += 1
                
        if table.merges:
            for merge in table.merges:
                try:
                    for r in range(row_offset + merge.startRow + 1, row_offset + merge.endRow + 2):
                        for c in range(merge.startCol + 1, merge.endCol + 2):
                            ws.cell(row=r, column=c).border = border_cell
                            
                    ws.merge_cells(
                        start_row=row_offset + merge.startRow + 1,
                        start_column=merge.startCol + 1,
                        end_row=row_offset + merge.endRow + 1,
                        end_column=merge.endCol + 1
                    )
                except Exception as e:
                    pass
                    
        current_row += 3
        
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.coordinate in ws.merged_cells:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(10, min(50, max_len + 3))
        
    return wb

def build_individual_zip_buffer(tables: List[ClientTableItem]) -> bytes:
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        used_filenames = set()
        for idx, table in enumerate(tables):
            wb = build_workbook_for_table_python(table, idx)
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            
            raw_file_name = sanitize_file_name(table.tableName, f"Bang_{idx + 1}")
            
            final_file_name = raw_file_name
            counter = 1
            while final_file_name.lower() in used_filenames:
                final_file_name = f"{raw_file_name} ({counter})"
                counter += 1
            used_filenames.add(final_file_name.lower())
            
            zip_file.writestr(f"{final_file_name}.xlsx", excel_buffer.getvalue())
            
    return zip_buffer.getvalue()

def build_multi_sheet_workbook(tables: List[ClientTableItem]) -> Workbook:
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
        
    used_sheet_names = set()
    
    for idx, table in enumerate(tables):
        sheet_name = sanitize_sheet_name(table.tableName, f"Sheet {idx + 1}")
        
        final_sheet_name = sheet_name
        counter = 1
        while final_sheet_name.lower() in used_sheet_names:
            suffix = f" ({counter})"
            max_len = 31 - len(suffix)
            final_sheet_name = sheet_name[:max_len] + suffix
            counter += 1
        used_sheet_names.add(final_sheet_name.lower())
        
        ws = wb.create_sheet(title=final_sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="4A6CF7", end_color="4A6CF7", fill_type="solid")
        font_body = Font(name="Arial", size=10, color="000000")
        fill_zebra = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
        
        thin_side = Side(border_style="thin", color="D0D0D0")
        medium_bottom_side = Side(border_style="medium", color="4A6CF7")
        border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=medium_bottom_side)
        
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        current_row = 1
        
        if table.headerRows is not None:
            for r_idx, r_data in enumerate(table.rows):
                is_header = r_idx in table.headerRows
                ws.row_dimensions[current_row].height = 25 if is_header else 20
                for c_idx, val in enumerate(r_data):
                    cell = ws.cell(row=current_row, column=c_idx + 1, value=val)
                    cell.alignment = align_center if is_header else align_left
                    if is_header:
                        cell.font = font_header
                        cell.fill = fill_header
                        cell.border = border_header
                    else:
                        cell.font = font_body
                        cell.border = border_cell
                        if r_idx % 2 == 1:
                            cell.fill = fill_zebra
                current_row += 1
        else:
            headers_list = []
            if table.headers:
                if isinstance(table.headers[0], list):
                    headers_list = table.headers
                else:
                    headers_list = [table.headers]
                    
            for h_row in headers_list:
                ws.row_dimensions[current_row].height = 25
                for c_idx, val in enumerate(h_row):
                    cell = ws.cell(row=current_row, column=c_idx + 1, value=val)
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center
                    cell.border = border_header
                current_row += 1
                
            for r_idx, r_data in enumerate(table.rows):
                ws.row_dimensions[current_row].height = 20
                is_zebra = (r_idx % 2 == 1)
                for c_idx, val in enumerate(r_data):
                    cell = ws.cell(row=current_row, column=c_idx + 1, value=val)
                    cell.font = font_body
                    if is_zebra:
                        cell.fill = fill_zebra
                    cell.alignment = align_left
                    cell.border = border_cell
                current_row += 1
                
        if table.merges:
            for merge in table.merges:
                try:
                    for r in range(merge.startRow + 1, merge.endRow + 2):
                        for c in range(merge.startCol + 1, merge.endCol + 2):
                            ws.cell(row=r, column=c).border = border_cell
                            
                    ws.merge_cells(
                        start_row=merge.startRow + 1,
                        start_column=merge.startCol + 1,
                        end_row=merge.endRow + 1,
                        end_column=merge.endCol + 1
                    )
                except Exception as e:
                    pass
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.coordinate in ws.merged_cells:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(10, min(50, max_len + 3))
            
    return wb

def export_tables(tables: List[ClientTableItem], options: ExportOptions) -> Tuple[bytes, str, str]:
    """Quyết định định dạng xuất file (ZIP, dọc gộp, hoặc nhiều Sheet) và trả về bytes dữ liệu cùng media_type và filename tương ứng."""
    if options.zip:
        zip_bytes = build_individual_zip_buffer(tables)
        return zip_bytes, "application/zip", "extracted_tables_individual.zip"
        
    elif options.verticalMerge:
        wb = build_combined_vertical_workbook(tables)
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        return excel_buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "extracted_data_merged.xlsx"
        
    else:
        wb = build_multi_sheet_workbook(tables)
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        return excel_buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "extracted_data_combined.xlsx"
